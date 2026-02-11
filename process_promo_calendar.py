"""
Promo Calendar Processor for Amazon Haul EU5 Forecasting Dashboard
Reads WW Haul Marketing.xlsx and generates promo regressors for the dashboard input file.

Usage:
    python process_promo_calendar.py <marketing_file> [--output <input_file>]
    
Example:
    python process_promo_calendar.py "WW Haul Marketing.xlsx" --output inputs_forecasting.xlsx
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import argparse
import sys
import os
import re


# === CONFIGURATION ===

# Volume impact levels to include (MEDIUM and above)
VOLUME_IMPACT_MAP = {
    'MEGA: >200% uplift': 3,
    'HIGH: 100%-200% uplift': 2,
    'MEDIUM: 50%-100% uplift': 1,
    'SMALL: 20%-50% uplift': 0,   # Excluded (below threshold)
    'MINIMAL: <20% uplift': 0,     # Excluded (below threshold)
}

VOLUME_IMPACT_MIN = 1  # Minimum numeric level to include (1=MEDIUM)

# Promo type re-encoding: Original file uses 1=HVE, 2=Dollar, 3=Discount
# We re-encode so higher = more impactful
PROMO_TYPE_RECODE = {
    1: 3,  # HVE (PD, PBDD, Peak) -> 3 (highest impact)
    2: 2,  # Dollar Deals -> 2
    3: 1,  # Discount Percentages -> 1 (lowest impact)
}

# Sheet name mapping
SHEET_MAP = {
    'UK': 'UK',
    'DE': 'DE',
    'FRITES': 'FRITES',  # Contains FR, IT, ES
}

# Marketplaces
MARKETPLACES = ['UK', 'DE', 'FR', 'IT', 'ES']


def parse_volume_impact(impact_str):
    """Parse volume impact string to numeric level.
    Returns 0 if below threshold or unparseable."""
    if pd.isna(impact_str) or not isinstance(impact_str, str):
        return 0
    
    impact_str = impact_str.strip()
    
    # Try exact match first
    if impact_str in VOLUME_IMPACT_MAP:
        return VOLUME_IMPACT_MAP[impact_str]
    
    # Try partial match
    impact_upper = impact_str.upper()
    if 'MEGA' in impact_upper:
        return 3
    elif 'HIGH' in impact_upper:
        return 2
    elif 'MEDIUM' in impact_upper:
        return 1
    elif 'SMALL' in impact_upper:
        return 0
    elif 'MINIMAL' in impact_upper:
        return 0
    
    return 0


def parse_promo_type(type_val):
    """Parse promo type value to integer."""
    if pd.isna(type_val):
        return None
    try:
        return int(float(type_val))
    except (ValueError, TypeError):
        return None


def parse_discount_pct(discount_val):
    """Parse discount percentage to integer. Returns 0 if not applicable."""
    if pd.isna(discount_val):
        return 0
    
    # Handle string values like "5, 10"
    if isinstance(discount_val, str):
        # Take the max value if comma-separated
        parts = discount_val.replace(',', ' ').split()
        max_val = 0
        for part in parts:
            try:
                max_val = max(max_val, int(float(part)))
            except (ValueError, TypeError):
                pass
        return max_val
    
    try:
        return int(float(discount_val))
    except (ValueError, TypeError):
        return 0


def get_iso_week_label(date):
    """Convert date to ISO week label (Wk## YYYY) matching dashboard format.
    Uses Monday-based ISO weeks."""
    if pd.isna(date):
        return None
    
    if isinstance(date, str):
        try:
            date = pd.to_datetime(date)
        except Exception:
            return None
    
    iso_cal = date.isocalendar()
    return f"Wk{iso_cal[1]:02d} {iso_cal[0]}"


def date_to_iso_weeks(start_date, end_date):
    """Get all ISO week labels that a date range touches."""
    if pd.isna(start_date) or pd.isna(end_date):
        return []
    
    try:
        start = pd.to_datetime(start_date)
        end = pd.to_datetime(end_date)
    except Exception:
        return []
    
    weeks = set()
    current = start
    while current <= end:
        label = get_iso_week_label(current)
        if label:
            weeks.add(label)
        current += timedelta(days=1)
    
    return sorted(weeks)


def read_marketing_sheet(file_path, sheet_name):
    """Read a marketing sheet and return structured promo events.
    
    Returns list of dicts with keys:
        marketplace, start_date, end_date, promo_type, discount_pct, 
        volume_impact_str, volume_impact_num, description
    """
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
    except Exception as e:
        print(f"  Warning: Could not read sheet '{sheet_name}': {e}")
        return []
    
    events = []
    
    # Expected columns (by position or name)
    # Col 0: Country Code
    # Col 1: start_date
    # Col 2: end_date
    # Col 3: Type of Promotion (header is long but column D)
    # Col 4: Discount Percentage
    # Col 5: Promo Description
    # Col 6: Channel
    # Col 7: Target Audience
    # Col 8: Total Volume Impact Level
    
    cols = df.columns.tolist()
    
    for idx, row in df.iterrows():
        # Get marketplace code
        mp = row.iloc[0] if len(cols) > 0 else None
        if pd.isna(mp) or str(mp).strip() not in MARKETPLACES:
            # Try to infer from sheet name
            if sheet_name == 'UK':
                mp = 'UK'
            elif sheet_name == 'DE':
                mp = 'DE'
            else:
                continue
        else:
            mp = str(mp).strip()
        
        # Get dates
        start_date = row.iloc[1] if len(cols) > 1 else None
        end_date = row.iloc[2] if len(cols) > 2 else None
        
        if pd.isna(start_date) or pd.isna(end_date):
            continue
        
        # Get promo type
        raw_type = row.iloc[3] if len(cols) > 3 else None
        promo_type = parse_promo_type(raw_type)
        if promo_type is None:
            continue
        
        # Get discount percentage (only for type 3)
        raw_discount = row.iloc[4] if len(cols) > 4 else None
        discount_pct = parse_discount_pct(raw_discount)
        
        # Get description
        description = str(row.iloc[5]).strip() if len(cols) > 5 and pd.notna(row.iloc[5]) else ''
        
        # Get volume impact
        raw_impact = row.iloc[8] if len(cols) > 8 else None
        volume_impact_str = str(raw_impact).strip() if pd.notna(raw_impact) else ''
        volume_impact_num = parse_volume_impact(raw_impact)
        
        events.append({
            'marketplace': mp,
            'start_date': pd.to_datetime(start_date),
            'end_date': pd.to_datetime(end_date),
            'promo_type_original': promo_type,
            'promo_type_encoded': PROMO_TYPE_RECODE.get(promo_type, 0),
            'discount_pct': discount_pct,
            'volume_impact_str': volume_impact_str,
            'volume_impact_num': volume_impact_num,
            'description': description,
        })
    
    return events


def aggregate_events_to_weeks(events):
    """Aggregate promo events into weekly regressors per marketplace.
    
    For each week and marketplace:
    - promo_type: From the highest volume impact promo, re-encoded (HVE=3, Dollar=2, Discount=1)
    - discount_pct: From the highest volume impact promo (0 if type != Discount%)
    - volume_impact: Max numeric impact across all promos (0=None, 1=MEDIUM, 2=HIGH, 3=MEGA)
    - promo_count: Count of all MEDIUM+ promos active that week
    
    Returns: dict[marketplace][week_label] = {promo_type, discount_pct, volume_impact, promo_count}
    """
    # Filter to MEDIUM+ events only
    filtered = [e for e in events if e['volume_impact_num'] >= VOLUME_IMPACT_MIN]
    
    print(f"\n  Total events: {len(events)}")
    print(f"  After MEDIUM+ filter: {len(filtered)}")
    
    # Build week-level data
    weekly_data = {}  # mp -> week -> list of events
    
    for event in filtered:
        mp = event['marketplace']
        weeks = date_to_iso_weeks(event['start_date'], event['end_date'])
        
        if mp not in weekly_data:
            weekly_data[mp] = {}
        
        for week in weeks:
            if week not in weekly_data[mp]:
                weekly_data[mp][week] = []
            weekly_data[mp][week].append(event)
    
    # Aggregate per week per marketplace
    result = {}
    
    for mp in MARKETPLACES:
        result[mp] = {}
        
        if mp not in weekly_data:
            continue
        
        for week, week_events in weekly_data[mp].items():
            # Sort by volume impact (descending) to find dominant promo
            sorted_events = sorted(week_events, key=lambda e: e['volume_impact_num'], reverse=True)
            dominant = sorted_events[0]
            
            # promo_type: from dominant (highest impact) event
            promo_type = dominant['promo_type_encoded']
            
            # discount_pct: from dominant event (only if original type was Discount %)
            discount_pct = dominant['discount_pct'] if dominant['promo_type_original'] == 3 else 0
            
            # volume_impact: max across all events
            volume_impact = max(e['volume_impact_num'] for e in week_events)
            
            # promo_count: number of MEDIUM+ events active
            promo_count = len(week_events)
            
            result[mp][week] = {
                'promo_type': promo_type,
                'discount_pct': discount_pct,
                'volume_impact': volume_impact,
                'promo_count': promo_count,
            }
    
    return result


def sort_week_labels(week_labels):
    """Sort week labels chronologically."""
    def week_sort_key(label):
        match = re.match(r'Wk(\d+)\s+(\d{4})', label)
        if match:
            return (int(match.group(2)), int(match.group(1)))
        return (9999, 99)
    
    return sorted(week_labels, key=week_sort_key)


def write_promo_regressors_sheet(output_file, weekly_regressors):
    """Write the Promo Regressors sheet to the Excel input file.
    
    Format:
    promo_type
    MP | Wk19 2025 | Wk20 2025 | ...
    UK |     3     |     0     | ...
    DE |     2     |     1     | ...
    ...
    
    discount_pct
    MP | Wk19 2025 | Wk20 2025 | ...
    UK |    50     |     0     | ...
    ...
    
    volume_impact
    MP | Wk19 2025 | Wk20 2025 | ...
    ...
    
    promo_count
    MP | Wk19 2025 | Wk20 2025 | ...
    ...
    """
    # Collect all unique weeks across all marketplaces
    all_weeks = set()
    for mp in MARKETPLACES:
        if mp in weekly_regressors:
            all_weeks.update(weekly_regressors[mp].keys())
    
    all_weeks = sort_week_labels(list(all_weeks))
    
    if not all_weeks:
        print("  No weeks to write!")
        return False
    
    print(f"\n  Writing {len(all_weeks)} weeks for {len(MARKETPLACES)} marketplaces")
    print(f"  Week range: {all_weeks[0]} → {all_weeks[-1]}")
    
    # Build the dataframe rows
    regressor_names = ['promo_type', 'discount_pct', 'volume_impact', 'promo_count']
    rows = []
    
    for regressor in regressor_names:
        # Section header row
        header_row = [regressor] + [''] * len(all_weeks)
        rows.append(header_row)
        
        # MP header row
        mp_header = ['MP'] + all_weeks
        rows.append(mp_header)
        
        # Data rows per marketplace
        for mp in MARKETPLACES:
            row = [mp]
            for week in all_weeks:
                if mp in weekly_regressors and week in weekly_regressors[mp]:
                    row.append(weekly_regressors[mp][week][regressor])
                else:
                    row.append(0)
            rows.append(row)
        
        # EU5 average row
        eu5_row = ['EU5']
        for week in all_weeks:
            values = []
            for mp in MARKETPLACES:
                if mp in weekly_regressors and week in weekly_regressors[mp]:
                    values.append(weekly_regressors[mp][week][regressor])
                else:
                    values.append(0)
            eu5_row.append(round(sum(values) / len(values), 2) if values else 0)
        rows.append(eu5_row)
        
        # Blank row between sections
        rows.append([''] * (len(all_weeks) + 1))
    
    # Create DataFrame
    max_cols = max(len(r) for r in rows)
    padded_rows = [r + [''] * (max_cols - len(r)) for r in rows]
    
    df_out = pd.DataFrame(padded_rows)
    
    # Write to Excel
    try:
        from openpyxl import load_workbook
        
        if os.path.exists(output_file):
            # Load existing workbook and add/replace sheet
            wb = load_workbook(output_file)
            
            # Remove old promo sheets
            for old_sheet in ['Promo Scores', 'Promo Regressors']:
                if old_sheet in wb.sheetnames:
                    del wb[old_sheet]
                    print(f"  Removed old '{old_sheet}' sheet")
            
            wb.save(output_file)
            
            # Write new sheet using openpyxl via pandas
            with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df_out.to_excel(writer, sheet_name='Promo Regressors', index=False, header=False)
        else:
            # Create new file
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df_out.to_excel(writer, sheet_name='Promo Regressors', index=False, header=False)
        
        print(f"\n  ✓ Wrote 'Promo Regressors' sheet to {output_file}")
        return True
        
    except Exception as e:
        print(f"\n  ✗ Error writing Excel: {e}")
        import traceback
        traceback.print_exc()
        return False


def print_summary(weekly_regressors):
    """Print a summary of the generated regressors."""
    print("\n" + "=" * 60)
    print("  PROMO REGRESSORS SUMMARY")
    print("=" * 60)
    
    for mp in MARKETPLACES:
        if mp not in weekly_regressors:
            print(f"\n  {mp}: No promo weeks")
            continue
        
        weeks = weekly_regressors[mp]
        n_weeks = len(weeks)
        
        # Count by volume impact
        mega_weeks = sum(1 for w in weeks.values() if w['volume_impact'] == 3)
        high_weeks = sum(1 for w in weeks.values() if w['volume_impact'] == 2)
        medium_weeks = sum(1 for w in weeks.values() if w['volume_impact'] == 1)
        
        # Multi-promo weeks
        multi_promo = sum(1 for w in weeks.values() if w['promo_count'] > 1)
        max_overlap = max((w['promo_count'] for w in weeks.values()), default=0)
        
        # Average discount
        discounts = [w['discount_pct'] for w in weeks.values() if w['discount_pct'] > 0]
        avg_discount = round(sum(discounts) / len(discounts), 1) if discounts else 0
        
        print(f"\n  {mp}: {n_weeks} promo weeks")
        print(f"    MEGA: {mega_weeks} | HIGH: {high_weeks} | MEDIUM: {medium_weeks}")
        print(f"    Multi-promo weeks: {multi_promo} (max overlap: {max_overlap})")
        print(f"    Avg discount (when applicable): {avg_discount}%")


def main():
    parser = argparse.ArgumentParser(description='Process promo calendar into dashboard regressors')
    parser.add_argument('marketing_file', help='Path to WW Haul Marketing.xlsx')
    parser.add_argument('--output', '-o', default='inputs_forecasting.xlsx',
                       help='Path to dashboard input file (default: inputs_forecasting.xlsx)')
    
    args = parser.parse_args()
    
    if not os.path.exists(args.marketing_file):
        print(f"Error: File not found: {args.marketing_file}")
        sys.exit(1)
    
    print("=" * 60)
    print("  PROMO CALENDAR PROCESSOR")
    print("=" * 60)
    print(f"\n  Input:  {args.marketing_file}")
    print(f"  Output: {args.output}")
    
    # Read all sheets
    all_events = []
    
    for sheet_key, sheet_name in SHEET_MAP.items():
        print(f"\n  Reading sheet: {sheet_name}")
        events = read_marketing_sheet(args.marketing_file, sheet_name)
        print(f"    Found {len(events)} events")
        all_events.extend(events)
    
    print(f"\n  Total events across all sheets: {len(all_events)}")
    
    # Aggregate to weekly regressors
    print("\n  Aggregating to weekly regressors...")
    weekly_regressors = aggregate_events_to_weeks(all_events)
    
    # Print summary
    print_summary(weekly_regressors)
    
    # Write output
    print(f"\n  Writing to {args.output}...")
    success = write_promo_regressors_sheet(args.output, weekly_regressors)
    
    if success:
        print("\n  ✓ Done! Promo regressors have been written to the input file.")
        print("    You can now upload the input file to the dashboard.")
    else:
        print("\n  ✗ Failed to write output file.")
        sys.exit(1)


if __name__ == '__main__':
    main()
