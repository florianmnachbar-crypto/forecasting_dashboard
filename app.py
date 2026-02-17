"""
Amazon Haul EU5 Forecasting Dashboard
Flask Application - Main Entry Point
"""

import os
import io
import csv
import json
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_from_directory, Response, make_response
from werkzeug.utils import secure_filename
from data_processor import DataProcessor
from forecaster import Forecaster

# Initialize Flask app
app = Flask(__name__)

# App version
APP_VERSION = "2.7.0"

# Configuration
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['ALLOWED_EXTENSIONS'] = {'xlsx', 'xls'}

# SharePoint inputs file path (auto-load on startup if accessible)
SHAREPOINT_INPUTS_FILE = r"\\share.amazon.com@SSL\DavWWWRoot\sites\eu-haul\Shared Documents\Forecasting\inputs_forecasting.xlsx"

# Local fallback file
LOCAL_INPUTS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'inputs_forecasting.xlsx')

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Global data processor instance
data_processor = None
current_file = None


def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']


@app.route('/')
def index():
    """Render the main dashboard page"""
    return render_template('dashboard.html')


@app.route('/api/upload', methods=['POST'])
def upload_file():
    """Handle file upload"""
    global data_processor, current_file
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file provided'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'success': False, 'error': 'Invalid file type. Please upload an Excel file (.xlsx or .xls)'}), 400
    
    try:
        # Save the file
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # Process the file
        data_processor = DataProcessor()
        data_processor._backtest_cache = {}  # Clear backtest cache on new upload
        success, message = data_processor.load_excel(filepath)
        
        if success:
            current_file = filename
            # Recalculate EU5 totals
            data_processor.calculate_eu5_totals()
            return jsonify({
                'success': True,
                'message': message,
                'filename': filename,
                'metrics': list(data_processor.data.keys()),
                'marketplaces': list(set(mp for metric in data_processor.data.values() for mp in metric.keys()))
            })
        else:
            return jsonify({'success': False, 'error': message}), 400
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/data', methods=['GET'])
def get_data():
    """Get all loaded data including manual forecast if available"""
    global data_processor
    
    if data_processor is None:
        return jsonify({'success': False, 'error': 'No data loaded. Please upload a file first.'}), 400
    
    try:
        data = data_processor.get_all_data()
        manual_forecast = data_processor.get_manual_forecast_data()
        
        response = {
            'success': True,
            'data': data,
            'metrics': DataProcessor.METRICS,
            'marketplaces': DataProcessor.MARKETPLACES,
            'has_manual_forecast': data_processor.has_manual_forecast
        }
        
        if manual_forecast:
            response['manual_forecast'] = manual_forecast
        
        return jsonify(response)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/accuracy', methods=['GET'])
def get_accuracy():
    """Get forecast accuracy metrics for manual forecast vs actuals
    
    Query parameters:
    - timeframe: 'total' (all overlap), 't4w' (last 4 weeks), or 'cw' (current week only)
    """
    global data_processor
    
    if data_processor is None:
        return jsonify({'success': False, 'error': 'No data loaded'}), 400
    
    if not data_processor.has_manual_forecast:
        return jsonify({
            'success': True,
            'has_manual_forecast': False,
            'accuracy': None
        })
    
    try:
        # Get timeframe from query parameter (default to 'total')
        timeframe = request.args.get('timeframe', 'total')
        # Validate timeframe
        if timeframe not in ['total', 't4w', 'cw']:
            timeframe = 'total'
        
        accuracy = data_processor.get_all_accuracy_metrics(timeframe=timeframe)
        return jsonify({
            'success': True,
            'has_manual_forecast': True,
            'accuracy': accuracy,
            'timeframe': timeframe
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/promo-scores', methods=['GET'])
def get_promo_scores():
    """Get promo scores data for visualization overlays.
    
    For regressor format: also includes volume_impact per week per MP
    for chart overlay coloring (historic + future).
    """
    global data_processor
    
    if data_processor is None:
        return jsonify({'success': False, 'error': 'No data loaded'}), 400
    
    try:
        if not data_processor.has_promo_scores:
            return jsonify({
                'success': True,
                'has_promo_scores': False,
                'promo_data': None
            })
        
        promo_data = data_processor.get_promo_scores_data()
        
        # If regressor format, also include the full regressor data
        if data_processor.promo_format == 'regressors' and data_processor.promo_regressors:
            promo_data['regressors'] = data_processor.promo_regressors
            promo_data['format'] = 'regressors'
            promo_data['volume_impact_labels'] = {
                0: 'No Promo', 1: 'MEDIUM', 2: 'HIGH', 3: 'MEGA'
            }
            promo_data['discount_values'] = data_processor.get_distinct_discount_values()
        
        return jsonify({
            'success': True,
            'has_promo_scores': True,
            'promo_data': promo_data
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/promo-analysis', methods=['GET'])
def get_promo_analysis():
    """Get promo uplift analysis for all metrics and marketplaces.
    
    Returns regressor-based analysis if 4-regressor format is loaded,
    otherwise falls back to legacy band-based analysis.
    """
    global data_processor
    
    if data_processor is None:
        return jsonify({'success': False, 'error': 'No data loaded'}), 400
    
    try:
        if not data_processor.has_promo_scores:
            return jsonify({
                'success': True,
                'has_promo_scores': False,
                'analysis': None
            })
        
        # Use regressor-based analysis if available
        if data_processor.promo_format == 'regressors':
            analysis = data_processor.get_all_regressor_analysis()
            return jsonify({
                'success': True,
                'has_promo_scores': True,
                'promo_format': 'regressors',
                'analysis': analysis
            })
        else:
            analysis = data_processor.get_all_promo_analysis()
            return jsonify({
                'success': True,
                'has_promo_scores': True,
                'promo_format': 'legacy',
                'analysis': analysis
            })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/forecast-uplift', methods=['GET'])
def get_forecast_uplift():
    """Get forecast data with promo uplift applied
    
    Note: Promo uplift is now ONLY applied to the SARIMAX model forecast, 
    not to manual forecasts. Manual forecasts are returned as-is.
    Use /api/forecast with include_promo=True for SARIMAX with promo regressor.
    """
    global data_processor
    
    if data_processor is None:
        return jsonify({'success': False, 'error': 'No data loaded'}), 400
    
    try:
        # Promo uplift is now only available via SARIMAX model, not manual forecast
        return jsonify({
            'success': True,
            'has_uplift_data': False,
            'uplift_data': None,
            'message': 'Promo uplift is now applied via SARIMAX model regressor. Use /api/forecast with include_promo=true to generate promo-adjusted model forecasts.'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/statistics', methods=['GET'])
def get_statistics():
    """Get summary statistics for all metrics and marketplaces"""
    global data_processor
    
    if data_processor is None:
        return jsonify({'success': False, 'error': 'No data loaded'}), 400
    
    try:
        stats = {}
        for metric in DataProcessor.METRICS:
            stats[metric] = {}
            for mp in DataProcessor.MARKETPLACES:
                stat = data_processor.get_summary_statistics(metric, mp)
                if stat:
                    stats[metric][mp] = stat
        
        return jsonify({
            'success': True,
            'statistics': stats
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/forecast', methods=['POST'])
def generate_forecast():
    """Generate forecast for specified metric and marketplace
    
    If promo scores are available and include_promo=True, applies promo uplift to SARIMAX model
    """
    global data_processor
    
    if data_processor is None:
        return jsonify({'success': False, 'error': 'No data loaded'}), 400
    
    try:
        params = request.get_json()
        metric = params.get('metric', 'Net Ordered Units')
        marketplace = params.get('marketplace', 'UK')
        model_type = params.get('model', 'sarimax')
        use_seasonality = params.get('seasonality', True)
        include_promo = params.get('include_promo', False)  # NEW: whether to include promo as regressor
        
        # Get the data
        df = data_processor.get_dataframe(metric, marketplace)
        
        if df is None or df.empty:
            return jsonify({
                'success': False,
                'error': f'No data available for {metric} - {marketplace}'
            }), 400
        
        # Generate forecast
        forecaster = Forecaster(forecast_horizon=12)
        
        # Prepare promo scores if requested and available
        exog = None
        future_exog = None
        promo_info = None
        
        if include_promo and data_processor.has_promo_scores and model_type.lower() == 'sarimax':
            exog, future_exog, promo_info = _prepare_promo_exog(data_processor, metric, marketplace, df, forecaster.forecast_horizon)
        
        if model_type.lower() == 'sarimax':
            forecast = forecaster.forecast_sarimax(df, use_seasonality=use_seasonality, exog=exog, future_exog=future_exog)
        else:
            forecast = forecaster.generate_forecast(df, model_type=model_type, use_seasonality=use_seasonality)
        
        if forecast is None:
            return jsonify({
                'success': False,
                'error': 'Failed to generate forecast. Insufficient data.'
            }), 400
        
        # Add promo info if used
        if promo_info:
            forecast['promo_info'] = promo_info
        
        # Calculate forecast statistics
        forecast_stats = {
            'total': round(sum(forecast['values']), 2),
            'average': round(sum(forecast['values']) / len(forecast['values']), 2),
            'min': round(min(forecast['values']), 2),
            'max': round(max(forecast['values']), 2)
        }
        
        return jsonify({
            'success': True,
            'forecast': forecast,
            'forecast_statistics': forecast_stats,
            'metric': metric,
            'marketplace': marketplace,
            'promo_uplift_applied': promo_info is not None
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


def _prepare_promo_exog(data_processor, metric, marketplace, df, forecast_horizon):
    """Prepare exogenous promo score data for SARIMAX model
    
    Returns (historical_exog, future_exog, promo_info) tuple
    """
    import pandas as pd
    from datetime import timedelta
    
    try:
        # Build historical promo scores aligned with actuals
        promo_scores = []
        weeks_with_scores = 0
        
        for _, row in df.iterrows():
            week_label = data_processor.format_week_label(row['ds'])
            score = data_processor.get_promo_score_for_week(marketplace, week_label)
            
            if score is not None:
                promo_scores.append(score)
                weeks_with_scores += 1
            else:
                # Use baseline score (1.0) for missing weeks
                promo_scores.append(1.0)
        
        # Create exog dataframe
        exog = pd.DataFrame({
            'ds': df['ds'],
            'promo_score': promo_scores
        })
        
        # Prepare future promo scores
        last_date = df['ds'].max()
        future_dates = [last_date + timedelta(weeks=i+1) for i in range(forecast_horizon)]
        future_scores = []
        
        for future_date in future_dates:
            week_label = data_processor.format_week_label(future_date)
            score = data_processor.get_promo_score_for_week(marketplace, week_label)
            future_scores.append(score if score is not None else 1.0)  # Default to baseline
        
        future_exog = pd.DataFrame({
            'ds': future_dates,
            'promo_score': future_scores
        })
        
        promo_info = {
            'historical_weeks_with_scores': weeks_with_scores,
            'total_historical_weeks': len(df),
            'future_scores': [{'week': data_processor.format_week_label(d), 'score': s} 
                             for d, s in zip(future_dates, future_scores)]
        }
        
        return exog, future_exog, promo_info
        
    except Exception as e:
        print(f"Warning: Could not prepare promo exog: {e}")
        return None, None, None


# Maximum Transit Conversion rate cap (10%)
MAX_TRANSIT_CONVERSION = 0.10

# UPO cap multiplier (MP historical max × 2)
UPO_CAP_MULTIPLIER = 2.0

# Transits cap multiplier (MP historical max × 3)
TRANSITS_CAP_MULTIPLIER = 3.0


def _get_historical_max(data_processor, metric, marketplace):
    """Get historical maximum value for a metric and marketplace"""
    try:
        df = data_processor.get_dataframe(metric, marketplace)
        if df is not None and not df.empty:
            return df['y'].max()
    except Exception:
        pass
    return None


def _cap_transits(forecast, mp_historical_max, eu5_historical_max):
    """Cap Transits forecasts to prevent unrealistic extrapolations
    
    Cap = min(EU5_historical_max, MP_historical_max × 3)
    
    This ensures:
    - No marketplace can forecast more transits than EU5 ever had
    - No marketplace can forecast more than 3× their own historical max
    """
    if forecast is None:
        return forecast
    
    if mp_historical_max is None and eu5_historical_max is None:
        return forecast  # No historical data to cap against
    
    # Calculate cap: min of EU5 max and MP max × 3
    caps = []
    if eu5_historical_max is not None:
        caps.append(eu5_historical_max)
    if mp_historical_max is not None:
        caps.append(mp_historical_max * TRANSITS_CAP_MULTIPLIER)
    
    transits_cap = min(caps) if caps else None
    
    if transits_cap is None:
        return forecast
    
    capped_count = 0
    
    # Cap values
    capped_values = []
    for v in forecast['values']:
        if v > transits_cap:
            capped_values.append(transits_cap)
            capped_count += 1
        else:
            capped_values.append(v)
    forecast['values'] = capped_values
    
    # Cap confidence intervals
    forecast['lower_bound'] = [min(v, transits_cap) for v in forecast['lower_bound']]
    forecast['upper_bound'] = [min(v, transits_cap) for v in forecast['upper_bound']]
    
    # Add cap info to model_info
    if 'model_info' not in forecast:
        forecast['model_info'] = {}
    if capped_count > 0:
        forecast['model_info']['transits_capped'] = True
        forecast['model_info']['transits_capped_weeks'] = capped_count
        forecast['model_info']['transits_cap_value'] = transits_cap
        # Update model name
        current_model = forecast.get('model', 'SARIMAX')
        if '(Capped)' not in current_model:
            forecast['model'] = current_model + ' (Capped)'
    
    return forecast


def _cap_upo(forecast, mp_historical_max):
    """Cap UPO forecasts to prevent unrealistic extrapolations
    
    Cap = MP_historical_max × 2
    """
    if forecast is None:
        return forecast
    
    if mp_historical_max is None:
        return forecast  # No historical data to cap against
    
    upo_cap = mp_historical_max * UPO_CAP_MULTIPLIER
    
    capped_count = 0
    
    # Cap values
    capped_values = []
    for v in forecast['values']:
        if v > upo_cap:
            capped_values.append(upo_cap)
            capped_count += 1
        else:
            capped_values.append(v)
    forecast['values'] = capped_values
    
    # Cap confidence intervals
    forecast['lower_bound'] = [min(v, upo_cap) for v in forecast['lower_bound']]
    forecast['upper_bound'] = [min(v, upo_cap) for v in forecast['upper_bound']]
    
    # Add cap info to model_info
    if 'model_info' not in forecast:
        forecast['model_info'] = {}
    if capped_count > 0:
        forecast['model_info']['upo_capped'] = True
        forecast['model_info']['upo_capped_weeks'] = capped_count
        forecast['model_info']['upo_cap_value'] = upo_cap
        # Update model name
        current_model = forecast.get('model', 'SARIMAX')
        if '(Capped)' not in current_model:
            forecast['model'] = current_model + ' (Capped)'
    
    return forecast


def _cap_transit_conversion(forecast):
    """Cap Transit Conversion forecasts at a maximum of 10% (0.10)
    
    This prevents unrealistic forecasts when the model extrapolates from
    recent spikes in conversion rates.
    """
    if forecast is None:
        return forecast
    
    capped_count = 0
    
    # Cap values
    capped_values = []
    for v in forecast['values']:
        if v > MAX_TRANSIT_CONVERSION:
            capped_values.append(MAX_TRANSIT_CONVERSION)
            capped_count += 1
        else:
            capped_values.append(v)
    forecast['values'] = capped_values
    
    # Cap lower bound
    forecast['lower_bound'] = [min(v, MAX_TRANSIT_CONVERSION) for v in forecast['lower_bound']]
    
    # Cap upper bound
    forecast['upper_bound'] = [min(v, MAX_TRANSIT_CONVERSION) for v in forecast['upper_bound']]
    
    # Add cap info to model_info
    if 'model_info' not in forecast:
        forecast['model_info'] = {}
    if capped_count > 0:
        forecast['model_info']['conversion_capped'] = True
        forecast['model_info']['conversion_capped_weeks'] = capped_count
        forecast['model_info']['max_conversion'] = MAX_TRANSIT_CONVERSION
        forecast['model'] = forecast.get('model', 'SARIMAX') + ' (Capped)'
    
    return forecast


def _apply_promo_floor(promo_forecast, baseline_forecast, future_scores):
    """Apply floor logic: promo score > 1 cannot decrease forecast below baseline
    
    For each future week:
    - If promo_score > 1.0: use max(baseline, promo_adjusted) - promo can only help
    - If promo_score = 1.0 (no promo): use baseline forecast
    - If promo_score < 1.0: use promo_adjusted (allowed to be lower for weak promo weeks)
    
    This ensures that marking a week as "promo" can only improve the forecast, never decrease it.
    """
    floored_values = []
    floored_lower = []
    floored_upper = []
    floor_applied_count = 0
    baseline_used_count = 0
    
    for i in range(len(promo_forecast['values'])):
        promo_val = promo_forecast['values'][i]
        baseline_val = baseline_forecast['values'][i]
        promo_score = future_scores[i] if i < len(future_scores) else 1.0
        
        # Apply logic based on promo score
        if promo_score > 1.0:
            # Promo week - promo cannot decrease forecast below baseline
            if promo_val < baseline_val:
                # Promo model decreased the forecast - use baseline instead
                floored_values.append(baseline_val)
                floor_applied_count += 1
            else:
                floored_values.append(promo_val)
        elif promo_score == 1.0:
            # No promo week - use baseline forecast
            floored_values.append(baseline_val)
            baseline_used_count += 1
        else:
            # Low promo (score < 1) - use promo forecast (allowed to be lower)
            floored_values.append(promo_val)
        
        # Apply same logic to confidence intervals
        promo_lower = promo_forecast['lower_bound'][i]
        baseline_lower = baseline_forecast['lower_bound'][i]
        promo_upper = promo_forecast['upper_bound'][i]
        baseline_upper = baseline_forecast['upper_bound'][i]
        
        if promo_score > 1.0:
            floored_lower.append(max(promo_lower, baseline_lower))
            floored_upper.append(max(promo_upper, baseline_upper))
        elif promo_score == 1.0:
            floored_lower.append(baseline_lower)
            floored_upper.append(baseline_upper)
        else:
            floored_lower.append(promo_lower)
            floored_upper.append(promo_upper)
    
    # Return modified forecast
    result = promo_forecast.copy()
    result['values'] = floored_values
    result['lower_bound'] = floored_lower
    result['upper_bound'] = floored_upper
    
    # Add floor info to model_info
    if 'model_info' not in result:
        result['model_info'] = {}
    result['model_info']['promo_floor_applied'] = True
    result['model_info']['floor_applied_weeks'] = floor_applied_count
    result['model_info']['baseline_used_weeks'] = baseline_used_count
    
    # Update model name to indicate flooring
    if floor_applied_count > 0 or baseline_used_count > 0:
        result['model'] = result.get('model', 'SARIMAX') + ' +Promo'
    
    return result


@app.route('/api/forecast/all', methods=['POST'])
def generate_all_forecasts():
    """Generate forecasts for all metrics and marketplaces
    
    Net Ordered Units is calculated as: Transits × Transit Conversion × UPO
    The other metrics (Transits, Transit Conversion, UPO) are forecasted independently
    
    If include_promo=True, promo scores are used as SARIMAX regressors for all driver metrics
    """
    global data_processor
    
    if data_processor is None:
        return jsonify({'success': False, 'error': 'No data loaded'}), 400
    
    try:
        params = request.get_json()
        model_type = params.get('model', 'sarimax')
        use_seasonality = params.get('seasonality', True)
        include_promo = params.get('include_promo', False)  # NEW: include promo as SARIMAX regressor
        
        forecasts = {}
        forecaster = Forecaster(forecast_horizon=12)
        promo_applied_count = 0
        
        # Driver metrics that are forecasted independently
        driver_metrics = ['Transits', 'Transit Conversion', 'UPO']
        
        # Pre-calculate EU5 historical max for Transits (used for all MP caps)
        eu5_transits_max = _get_historical_max(data_processor, 'Transits', 'EU5')
        
        # First, forecast the driver metrics
        for metric in driver_metrics:
            forecasts[metric] = {}
            for mp in DataProcessor.MARKETPLACES:
                df = data_processor.get_dataframe(metric, mp)
                
                if df is not None and not df.empty and len(df) >= 4:
                    # Prepare promo exog if requested
                    exog = None
                    future_exog = None
                    promo_info = None
                    future_scores = None
                    
                    if include_promo and data_processor.has_promo_scores and model_type.lower() == 'sarimax':
                        exog, future_exog, promo_info = _prepare_promo_exog(
                            data_processor, metric, mp, df, forecaster.forecast_horizon
                        )
                        if promo_info:
                            promo_applied_count += 1
                            # Extract future promo scores for floor logic
                            future_scores = [item['score'] for item in promo_info.get('future_scores', [])]
                    
                    if model_type.lower() == 'sarimax':
                        # Generate baseline forecast (no promo) for floor comparison
                        baseline_forecast = None
                        if include_promo and exog is not None:
                            baseline_forecast = forecaster.forecast_sarimax(
                                df, use_seasonality=use_seasonality, exog=None, future_exog=None
                            )
                        
                        # Generate promo-adjusted forecast
                        forecast = forecaster.forecast_sarimax(
                            df, use_seasonality=use_seasonality, exog=exog, future_exog=future_exog
                        )
                        
                        # Apply floor: promo score > 1 cannot decrease forecast below baseline
                        if forecast and baseline_forecast and future_scores:
                            forecast = _apply_promo_floor(forecast, baseline_forecast, future_scores)
                    else:
                        forecast = forecaster.generate_forecast(
                            df, model_type=model_type, use_seasonality=use_seasonality
                        )
                    
                    if forecast:
                        # Apply caps to prevent unrealistic extrapolations
                        if metric == 'Transit Conversion':
                            forecast = _cap_transit_conversion(forecast)
                        elif metric == 'Transits':
                            mp_transits_max = _get_historical_max(data_processor, 'Transits', mp)
                            forecast = _cap_transits(forecast, mp_transits_max, eu5_transits_max)
                        elif metric == 'UPO':
                            mp_upo_max = _get_historical_max(data_processor, 'UPO', mp)
                            forecast = _cap_upo(forecast, mp_upo_max)
                        
                        if promo_info:
                            forecast['promo_info'] = promo_info
                        forecasts[metric][mp] = forecast
        
        # Calculate Net Ordered Units - either derived from drivers or direct forecast
        forecasts['Net Ordered Units'] = {}
        
        for mp in DataProcessor.MARKETPLACES:
            # Check if we have all driver forecasts for this marketplace
            has_transits = mp in forecasts.get('Transits', {})
            has_conversion = mp in forecasts.get('Transit Conversion', {})
            has_upo = mp in forecasts.get('UPO', {})
            
            if has_transits and has_conversion and has_upo:
                # Derive NOU from drivers: NOU = Transits × Conversion × UPO
                transits_fc = forecasts['Transits'][mp]
                conversion_fc = forecasts['Transit Conversion'][mp]
                upo_fc = forecasts['UPO'][mp]
                
                nou_values = []
                nou_lower = []
                nou_upper = []
                
                for i in range(len(transits_fc['values'])):
                    # Point estimate: product of means
                    t = transits_fc['values'][i]
                    c = conversion_fc['values'][i]
                    u = upo_fc['values'][i]
                    nou = t * c * u
                    nou_values.append(max(0, nou))
                    
                    # Confidence intervals using error propagation
                    t_low = transits_fc['lower_bound'][i]
                    c_low = conversion_fc['lower_bound'][i]
                    u_low = upo_fc['lower_bound'][i]
                    nou_lower.append(max(0, t_low * c_low * u_low))
                    
                    t_up = transits_fc['upper_bound'][i]
                    c_up = conversion_fc['upper_bound'][i]
                    u_up = upo_fc['upper_bound'][i]
                    nou_upper.append(max(0, t_up * c_up * u_up))
                
                forecasts['Net Ordered Units'][mp] = {
                    'dates': transits_fc['dates'],
                    'values': nou_values,
                    'lower_bound': nou_lower,
                    'upper_bound': nou_upper,
                    'model': 'Calculated (T×C×U)',
                    'model_info': {
                        'method': 'derived',
                        'formula': 'Transits × Transit Conversion × UPO',
                        'source_models': {
                            'Transits': transits_fc.get('model', 'SARIMAX'),
                            'Transit Conversion': conversion_fc.get('model', 'SARIMAX'),
                            'UPO': upo_fc.get('model', 'SARIMAX')
                        }
                    }
                }
            else:
                # FALLBACK: Forecast NOU directly when driver metrics unavailable
                df = data_processor.get_dataframe('Net Ordered Units', mp)
                
                if df is not None and not df.empty and len(df) >= 4:
                    # Prepare promo exog if requested
                    exog = None
                    future_exog = None
                    promo_info = None
                    future_scores = None
                    
                    if include_promo and data_processor.has_promo_scores and model_type.lower() == 'sarimax':
                        exog, future_exog, promo_info = _prepare_promo_exog(
                            data_processor, 'Net Ordered Units', mp, df, forecaster.forecast_horizon
                        )
                        if promo_info:
                            future_scores = [item['score'] for item in promo_info.get('future_scores', [])]
                    
                    if model_type.lower() == 'sarimax':
                        # Generate baseline forecast for floor comparison
                        baseline_forecast = None
                        if include_promo and exog is not None:
                            baseline_forecast = forecaster.forecast_sarimax(
                                df, use_seasonality=use_seasonality, exog=None, future_exog=None
                            )
                        
                        forecast = forecaster.forecast_sarimax(
                            df, use_seasonality=use_seasonality, exog=exog, future_exog=future_exog
                        )
                        
                        if forecast and baseline_forecast and future_scores:
                            forecast = _apply_promo_floor(forecast, baseline_forecast, future_scores)
                    else:
                        forecast = forecaster.generate_forecast(
                            df, model_type=model_type, use_seasonality=use_seasonality
                        )
                    
                    if forecast:
                        # Update model info to indicate direct forecasting
                        forecast['model_info'] = {
                            'method': 'direct',
                            'reason': 'Driver metrics unavailable',
                            'missing_drivers': [
                                m for m in ['Transits', 'Transit Conversion', 'UPO']
                                if mp not in forecasts.get(m, {})
                            ]
                        }
                        forecast['model'] = forecast.get('model', 'SARIMAX') + ' (Direct)'
                        
                        if promo_info:
                            forecast['promo_info'] = promo_info
                        
                        forecasts['Net Ordered Units'][mp] = forecast
        
        return jsonify({
            'success': True,
            'forecasts': forecasts,
            'model': model_type,
            'seasonality': use_seasonality,
            'derived_metrics': ['Net Ordered Units']
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/latest-week', methods=['GET'])
def get_latest_week():
    """Get latest week overview with all metrics and marketplaces"""
    global data_processor
    
    if data_processor is None:
        return jsonify({'success': False, 'error': 'No data loaded'}), 400
    
    try:
        overview = data_processor.get_latest_week_overview()
        if overview is None:
            return jsonify({
                'success': False,
                'error': 'Could not calculate latest week overview'
            }), 400
        
        # Include promo assessment if available
        promo_assessment = None
        if data_processor.has_promo_scores and data_processor.promo_format == 'regressors':
            promo_assessment = data_processor.get_latest_week_promo_assessment()
        
        return jsonify({
            'success': True,
            'overview': overview,
            'has_manual_forecast': data_processor.has_manual_forecast,
            'promo_assessment': promo_assessment
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/status', methods=['GET'])
def get_status():
    """Get current application status"""
    global data_processor, current_file
    
    return jsonify({
        'success': True,
        'data_loaded': data_processor is not None,
        'current_file': current_file,
        'metrics': DataProcessor.METRICS,
        'marketplaces': DataProcessor.MARKETPLACES,
        'has_manual_forecast': data_processor.has_manual_forecast if data_processor else False
    })


@app.route('/api/historic-deviations', methods=['GET'])
def get_historic_deviations():
    """Get historic deviations for actuals vs manual forecast and model forecast
    
    Query parameters:
    - metric: Metric name (default: 'Net Ordered Units')
    - marketplace: Marketplace code (default: 'UK')
    - include_model: Whether to run model backtest (default: 'true')
    """
    global data_processor
    
    if data_processor is None:
        return jsonify({'success': False, 'error': 'No data loaded'}), 400
    
    try:
        metric = request.args.get('metric', 'Net Ordered Units')
        marketplace = request.args.get('marketplace', 'UK')
        include_model = request.args.get('include_model', 'true').lower() == 'true'
        
        # Get actuals
        actuals_df = data_processor.get_dataframe(metric, marketplace, is_forecast=False)
        if actuals_df is None or actuals_df.empty:
            return jsonify({
                'success': False,
                'error': f'No actuals data for {metric} - {marketplace}'
            }), 400
        
        # Build the response data
        deviations = []
        
        # Get manual forecast if available
        manual_df = None
        if data_processor.has_manual_forecast:
            manual_df = data_processor.get_dataframe(metric, marketplace, is_forecast=True)
        
        # Get model backtest forecasts if requested
        model_forecasts = None
        if include_model:
            try:
                model_forecasts = data_processor.generate_historic_model_forecasts(metric, marketplace)
            except Exception as e:
                print(f"Model backtest warning: {e}")
                model_forecasts = None
        
        # Process each week in actuals
        for idx, row in actuals_df.iterrows():
            week = row['week']
            date = row['ds']
            actual = row['y']
            date_str = date.strftime('%Y-%m-%d')
            
            record = {
                'week': week,
                'date': date_str,
                'actual': actual,
                'manual_forecast': None,
                'manual_dev': None,
                'manual_dev_pct': None,
                'model_forecast': None,
                'model_dev': None,
                'model_dev_pct': None
            }
            
            # Get manual forecast value for this week
            if manual_df is not None and not manual_df.empty:
                manual_match = manual_df[manual_df['ds'] == date]
                if not manual_match.empty:
                    manual_val = manual_match['y'].iloc[0]
                    record['manual_forecast'] = manual_val
                    if manual_val != 0:
                        dev = actual - manual_val
                        dev_pct = (dev / manual_val) * 100
                        record['manual_dev'] = round(dev, 4)
                        record['manual_dev_pct'] = round(dev_pct, 1)
            
            # Get model backtest forecast value for this week
            if model_forecasts is not None and date_str in model_forecasts:
                model_val = model_forecasts[date_str]
                record['model_forecast'] = round(model_val, 4)
                if model_val != 0:
                    dev = actual - model_val
                    dev_pct = (dev / model_val) * 100
                    record['model_dev'] = round(dev, 4)
                    record['model_dev_pct'] = round(dev_pct, 1)
            
            deviations.append(record)
        
        # Calculate summary stats using trailing 6 weeks (T6W) for accuracy comparison
        # Sort deviations by date descending, take most recent 6 with data
        sorted_devs = sorted(deviations, key=lambda d: d['date'], reverse=True)
        
        # T6W for manual FC: most recent 6 weeks where manual forecast exists
        t6w_manual = [d for d in sorted_devs if d['manual_dev_pct'] is not None][:6]
        # T6W for model FC: most recent 6 weeks where model forecast exists
        t6w_model = [d for d in sorted_devs if d['model_dev_pct'] is not None][:6]
        
        # All-time counts for display
        manual_devs = [d['manual_dev_pct'] for d in deviations if d['manual_dev_pct'] is not None]
        model_devs = [d['model_dev_pct'] for d in deviations if d['model_dev_pct'] is not None]
        
        # Calculate WMAPE for manual FC (T6W)
        manual_wmape = None
        if t6w_manual:
            manual_abs_errors = [abs(d['actual'] - d['manual_forecast']) for d in t6w_manual]
            manual_actuals = [d['actual'] for d in t6w_manual]
            total_manual_actual = sum(manual_actuals)
            if total_manual_actual > 0:
                manual_wmape = round(sum(manual_abs_errors) / total_manual_actual * 100, 1)
        
        # Calculate WMAPE for model FC (T6W)
        model_wmape = None
        if t6w_model:
            model_abs_errors = [abs(d['actual'] - d['model_forecast']) for d in t6w_model]
            model_actuals = [d['actual'] for d in t6w_model]
            total_model_actual = sum(model_actuals)
            if total_model_actual > 0:
                model_wmape = round(sum(model_abs_errors) / total_model_actual * 100, 1)
        
        # T6W bias
        manual_t6w_devs = [d['manual_dev_pct'] for d in t6w_manual]
        model_t6w_devs = [d['model_dev_pct'] for d in t6w_model]
        
        summary = {
            'total_weeks': len(deviations),
            'manual_forecast_weeks': len(t6w_manual),
            'manual_avg_dev': round(sum(manual_t6w_devs) / len(manual_t6w_devs), 1) if manual_t6w_devs else None,
            'manual_avg_abs_dev': round(sum(abs(d) for d in manual_t6w_devs) / len(manual_t6w_devs), 1) if manual_t6w_devs else None,
            'manual_wmape': manual_wmape,
            'manual_accuracy': round(100 - manual_wmape, 1) if manual_wmape is not None else None,
            'model_forecast_weeks': len(t6w_model),
            'model_avg_dev': round(sum(model_t6w_devs) / len(model_t6w_devs), 1) if model_t6w_devs else None,
            'model_avg_abs_dev': round(sum(abs(d) for d in model_t6w_devs) / len(model_t6w_devs), 1) if model_t6w_devs else None,
            'model_wmape': model_wmape,
            'model_accuracy': round(100 - model_wmape, 1) if model_wmape is not None else None,
            'timeframe': 'T6W',
        }
        
        return jsonify({
            'success': True,
            'metric': metric,
            'marketplace': marketplace,
            'deviations': deviations,
            'summary': summary,
            'has_manual_forecast': data_processor.has_manual_forecast,
            'has_model_backtest': model_forecasts is not None and len(model_devs) > 0
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/export/csv', methods=['GET'])
def export_csv():
    """Export all data as CSV file"""
    global data_processor
    
    if data_processor is None:
        return jsonify({'success': False, 'error': 'No data loaded'}), 400
    
    try:
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Section 1: Latest Week Overview
        writer.writerow(['=' * 50])
        writer.writerow(['LATEST WEEK OVERVIEW'])
        writer.writerow(['=' * 50])
        writer.writerow([])
        
        overview = data_processor.get_latest_week_overview()
        if overview and 'latest_week' in overview:
            writer.writerow(['Latest Week:', overview['latest_week']])
            writer.writerow([])
            
            # Header row
            header = ['Marketplace']
            for metric in DataProcessor.METRICS:
                header.extend([f'{metric} - Actual', f'{metric} - Forecast', f'{metric} - Deviation %'])
            writer.writerow(header)
            
            # Data rows
            for mp in ['EU5', 'UK', 'DE', 'FR', 'IT', 'ES']:
                if mp in overview.get('data', {}):
                    row = [mp]
                    for metric in DataProcessor.METRICS:
                        mp_data = overview['data'][mp].get(metric, {})
                        actual = mp_data.get('actual', '')
                        forecast = mp_data.get('manual_forecast', '')
                        deviation = mp_data.get('manual_dev_pct', '')
                        
                        # Format values
                        if actual != '' and actual is not None:
                            if metric in ['Transit Conversion']:
                                actual = f'{actual:.4f}' if isinstance(actual, (int, float)) else actual
                            elif metric in ['UPO']:
                                actual = f'{actual:.2f}' if isinstance(actual, (int, float)) else actual
                            else:
                                actual = f'{actual:,.0f}' if isinstance(actual, (int, float)) else actual
                        
                        if forecast != '' and forecast is not None:
                            if metric in ['Transit Conversion']:
                                forecast = f'{forecast:.4f}' if isinstance(forecast, (int, float)) else forecast
                            elif metric in ['UPO']:
                                forecast = f'{forecast:.2f}' if isinstance(forecast, (int, float)) else forecast
                            else:
                                forecast = f'{forecast:,.0f}' if isinstance(forecast, (int, float)) else forecast
                        
                        if deviation != '' and deviation is not None:
                            deviation = f'{deviation:.1f}%' if isinstance(deviation, (int, float)) else deviation
                        
                        row.extend([actual, forecast, deviation])
                    writer.writerow(row)
        
        writer.writerow([])
        writer.writerow([])
        
        # Section 2: Statistics Summary
        writer.writerow(['=' * 50])
        writer.writerow(['STATISTICS SUMMARY'])
        writer.writerow(['=' * 50])
        writer.writerow([])
        
        for metric in DataProcessor.METRICS:
            writer.writerow([f'--- {metric} ---'])
            writer.writerow(['Marketplace', 'Total', 'Average', 'Min', 'Max', 'T4W Total', 'T4W Average'])
            
            for mp in DataProcessor.MARKETPLACES:
                stats = data_processor.get_summary_statistics(metric, mp)
                if stats:
                    writer.writerow([
                        mp,
                        f'{stats.get("total", 0):,.2f}',
                        f'{stats.get("average", 0):,.2f}',
                        f'{stats.get("min", 0):,.2f}',
                        f'{stats.get("max", 0):,.2f}',
                        f'{stats.get("t4w_total", 0):,.2f}',
                        f'{stats.get("t4w_avg", 0):,.2f}'
                    ])
            writer.writerow([])
        
        writer.writerow([])
        
        # Section 3: Historical Data
        writer.writerow(['=' * 50])
        writer.writerow(['HISTORICAL DATA'])
        writer.writerow(['=' * 50])
        writer.writerow([])
        
        for metric in DataProcessor.METRICS:
            writer.writerow([f'--- {metric} ---'])
            
            # Get all weeks from UK or first available marketplace
            all_data = data_processor.get_all_data()
            if metric in all_data:
                metric_data = all_data[metric]
                
                # Collect all unique weeks
                all_weeks = set()
                for mp, mp_data in metric_data.items():
                    if isinstance(mp_data, dict) and 'weeks' in mp_data:
                        all_weeks.update(mp_data['weeks'])
                
                all_weeks = sorted(list(all_weeks))
                
                if all_weeks:
                    # Header
                    header = ['Week'] + DataProcessor.MARKETPLACES
                    writer.writerow(header)
                    
                    # Data rows
                    for week in all_weeks:
                        row = [week]
                        for mp in DataProcessor.MARKETPLACES:
                            value = ''
                            if mp in metric_data and isinstance(metric_data[mp], dict):
                                weeks = metric_data[mp].get('weeks', [])
                                values = metric_data[mp].get('values', [])
                                if week in weeks:
                                    idx = weeks.index(week)
                                    if idx < len(values):
                                        value = values[idx]
                            row.append(value if value != '' else '')
                        writer.writerow(row)
            
            writer.writerow([])
        
        # Create response
        output.seek(0)
        timestamp = datetime.now().strftime('%Y-%m-%d')
        filename = f'amazon_haul_eu5_export_{timestamp}.csv'
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/export/excel', methods=['GET'])
def export_excel():
    """Export all data as Excel file with pivoted MP×Week layout.
    
    Sheets:
    1. Latest Week Overview (compact)
    2. Statistics
    3-6. One per metric: rows = MP grouped by data series, columns = weeks
         Data series: Actuals, Manual FC, Model FC, Model FC +Promo
    """
    global data_processor
    
    if data_processor is None:
        return jsonify({'success': False, 'error': 'No data loaded'}), 400
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import timedelta
        
        wb = openpyxl.Workbook()
        
        # Styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='232F3E', end_color='232F3E', fill_type='solid')
        section_fills = {
            'Actuals': PatternFill(start_color='1B5E20', end_color='1B5E20', fill_type='solid'),
            'Manual FC': PatternFill(start_color='4A148C', end_color='4A148C', fill_type='solid'),
            'Model FC': PatternFill(start_color='0D47A1', end_color='0D47A1', fill_type='solid'),
            'Model FC +Promo': PatternFill(start_color='BF360C', end_color='BF360C', fill_type='solid'),
        }
        good_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        warn_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        bad_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        section_font = Font(bold=True, color='FFFFFF', size=11)
        
        mp_order = ['EU5', 'UK', 'DE', 'FR', 'IT', 'ES']
        
        # --- Generate model forecasts (baseline + promo) for export ---
        forecaster = Forecaster(forecast_horizon=12)
        model_forecasts = {}       # metric -> mp -> {dates, values}
        model_promo_forecasts = {}  # metric -> mp -> {dates, values}
        
        driver_metrics = ['Transits', 'Transit Conversion', 'UPO']
        eu5_transits_max = _get_historical_max(data_processor, 'Transits', 'EU5')
        
        # Generate baseline and promo forecasts for driver metrics
        for metric in driver_metrics:
            model_forecasts[metric] = {}
            model_promo_forecasts[metric] = {}
            
            for mp in DataProcessor.MARKETPLACES:
                df = data_processor.get_dataframe(metric, mp)
                if df is None or df.empty or len(df) < 4:
                    continue
                
                # Baseline forecast (no promo)
                fc_base = forecaster.forecast_sarimax(df, use_seasonality=True, exog=None, future_exog=None)
                if fc_base:
                    # Apply caps
                    if metric == 'Transit Conversion':
                        fc_base = _cap_transit_conversion(fc_base)
                    elif metric == 'Transits':
                        mp_max = _get_historical_max(data_processor, 'Transits', mp)
                        fc_base = _cap_transits(fc_base, mp_max, eu5_transits_max)
                    elif metric == 'UPO':
                        mp_max = _get_historical_max(data_processor, 'UPO', mp)
                        fc_base = _cap_upo(fc_base, mp_max)
                    model_forecasts[metric][mp] = fc_base
                
                # Promo forecast (if promo data available)
                if data_processor.has_promo_scores:
                    exog, future_exog, promo_info = _prepare_promo_exog(
                        data_processor, metric, mp, df, forecaster.forecast_horizon
                    )
                    if exog is not None:
                        fc_promo = forecaster.forecast_sarimax(df, use_seasonality=True, exog=exog, future_exog=future_exog)
                        if fc_promo and fc_base and promo_info:
                            future_scores = [item['score'] for item in promo_info.get('future_scores', [])]
                            fc_promo = _apply_promo_floor(fc_promo, fc_base, future_scores)
                            # Apply caps
                            if metric == 'Transit Conversion':
                                fc_promo = _cap_transit_conversion(fc_promo)
                            elif metric == 'Transits':
                                mp_max = _get_historical_max(data_processor, 'Transits', mp)
                                fc_promo = _cap_transits(fc_promo, mp_max, eu5_transits_max)
                            elif metric == 'UPO':
                                mp_max = _get_historical_max(data_processor, 'UPO', mp)
                                fc_promo = _cap_upo(fc_promo, mp_max)
                            model_promo_forecasts[metric][mp] = fc_promo
        
        # Derive Net Ordered Units from driver forecasts
        for fc_dict in [model_forecasts, model_promo_forecasts]:
            fc_dict['Net Ordered Units'] = {}
            for mp in DataProcessor.MARKETPLACES:
                t = fc_dict.get('Transits', {}).get(mp)
                c = fc_dict.get('Transit Conversion', {}).get(mp)
                u = fc_dict.get('UPO', {}).get(mp)
                if t and c and u:
                    nou_vals = [t['values'][i] * c['values'][i] * u['values'][i] for i in range(len(t['values']))]
                    fc_dict['Net Ordered Units'][mp] = {
                        'dates': t['dates'],
                        'values': [max(0, v) for v in nou_vals]
                    }
        
        # Helper: convert forecast date strings to week labels
        def dates_to_weeks(date_strs):
            weeks = []
            for ds in date_strs:
                d = datetime.strptime(ds, '%Y-%m-%d')
                wl = data_processor.format_week_label(d)
                weeks.append(wl)
            return weeks
        
        # --- Sheet 1: Latest Week Overview ---
        ws1 = wb.active
        ws1.title = 'Latest Week Overview'
        
        overview = data_processor.get_latest_week_overview()
        if overview and 'latest_week' in overview:
            ws1['A1'] = f'Latest Week: {overview["latest_week"]}'
            ws1['A1'].font = Font(bold=True, size=14)
            
            row = 3
            headers = ['Marketplace']
            for metric in DataProcessor.METRICS:
                headers.extend([f'{metric}\nActual', f'{metric}\nForecast', f'{metric}\nDev %'])
            
            for col, header in enumerate(headers, 1):
                cell = ws1.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
                ws1.column_dimensions[get_column_letter(col)].width = 12
            ws1.row_dimensions[row].height = 40
            
            for mp in mp_order:
                if mp in overview.get('data', {}):
                    row += 1
                    col = 1
                    ws1.cell(row=row, column=col, value=mp).border = thin_border
                    for metric in DataProcessor.METRICS:
                        mp_data = overview['data'][mp].get(metric, {})
                        actual = mp_data.get('actual')
                        forecast = mp_data.get('manual_forecast')
                        deviation = mp_data.get('manual_dev_pct')
                        col += 1
                        ws1.cell(row=row, column=col, value=actual if actual is not None else '').border = thin_border
                        col += 1
                        ws1.cell(row=row, column=col, value=forecast if forecast is not None else '').border = thin_border
                        col += 1
                        cell = ws1.cell(row=row, column=col, value=f'{deviation:.1f}%' if deviation is not None else '')
                        cell.border = thin_border
                        if deviation is not None:
                            abs_dev = abs(deviation)
                            if abs_dev < 20:
                                cell.fill = good_fill
                            elif abs_dev < 30:
                                cell.fill = warn_fill
                            else:
                                cell.fill = bad_fill
        
        # --- Sheet 2: Statistics ---
        ws2 = wb.create_sheet('Statistics')
        row = 1
        for metric in DataProcessor.METRICS:
            ws2.cell(row=row, column=1, value=metric).font = Font(bold=True, size=12)
            row += 1
            for col, h in enumerate(['Marketplace', 'Total', 'Average', 'Min', 'Max', 'Std Dev', 'Weeks'], 1):
                cell = ws2.cell(row=row, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                ws2.column_dimensions[get_column_letter(col)].width = 14
            row += 1
            for mp in DataProcessor.MARKETPLACES:
                stats = data_processor.get_summary_statistics(metric, mp)
                if stats:
                    ws2.cell(row=row, column=1, value=mp).border = thin_border
                    ws2.cell(row=row, column=2, value=stats.get('total', 0)).border = thin_border
                    ws2.cell(row=row, column=3, value=stats.get('average', 0)).border = thin_border
                    ws2.cell(row=row, column=4, value=stats.get('min', 0)).border = thin_border
                    ws2.cell(row=row, column=5, value=stats.get('max', 0)).border = thin_border
                    ws2.cell(row=row, column=6, value=stats.get('std_dev', 0)).border = thin_border
                    ws2.cell(row=row, column=7, value=stats.get('count', 0)).border = thin_border
                    row += 1
            row += 2
        
        # --- Sheets 3-6: Pivoted MP×Week per metric ---
        all_data = data_processor.get_all_data()
        manual_fc_data = data_processor.get_manual_forecast_data() if data_processor.has_manual_forecast else None
        
        for metric in DataProcessor.METRICS:
            sheet_name = metric[:31]
            ws = wb.create_sheet(sheet_name)
            
            # Collect all weeks across all series
            all_weeks = set()
            
            # From actuals
            if metric in all_data:
                for mp, mp_data in all_data[metric].items():
                    if isinstance(mp_data, dict) and 'weeks' in mp_data:
                        all_weeks.update(mp_data['weeks'])
            
            # From manual forecast
            if manual_fc_data and metric in manual_fc_data:
                for mp, mp_data in manual_fc_data[metric].items():
                    if isinstance(mp_data, dict) and 'weeks' in mp_data:
                        all_weeks.update(mp_data['weeks'])
            
            # From model forecast
            if metric in model_forecasts:
                for mp, fc in model_forecasts[metric].items():
                    if fc and 'dates' in fc:
                        all_weeks.update(dates_to_weeks(fc['dates']))
            
            # From model forecast +promo
            if metric in model_promo_forecasts:
                for mp, fc in model_promo_forecasts[metric].items():
                    if fc and 'dates' in fc:
                        all_weeks.update(dates_to_weeks(fc['dates']))
            
            all_weeks = sorted(list(all_weeks))
            if not all_weeks:
                continue
            
            week_col_map = {w: idx + 2 for idx, w in enumerate(all_weeks)}  # col 1 = label, col 2+ = weeks
            
            # Write week header row
            ws.cell(row=1, column=1, value='Series / MP').font = Font(bold=True)
            ws.column_dimensions['A'].width = 18
            for w, col_idx in week_col_map.items():
                cell = ws.cell(row=1, column=col_idx, value=w)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
                ws.column_dimensions[get_column_letter(col_idx)].width = 12
            
            current_row = 2
            
            # Helper to write a section
            def write_section(section_name, data_getter, fill):
                nonlocal current_row
                # Section header row
                cell = ws.cell(row=current_row, column=1, value=section_name)
                cell.font = section_font
                cell.fill = fill
                # Fill the header across all week columns
                for col_idx in range(2, 2 + len(all_weeks)):
                    ws.cell(row=current_row, column=col_idx).fill = fill
                    ws.cell(row=current_row, column=col_idx).border = thin_border
                cell.border = thin_border
                current_row += 1
                
                # MP rows
                for mp in mp_order:
                    values_by_week = data_getter(mp)
                    if not values_by_week:
                        continue
                    
                    ws.cell(row=current_row, column=1, value=mp).border = thin_border
                    for w, col_idx in week_col_map.items():
                        val = values_by_week.get(w)
                        cell = ws.cell(row=current_row, column=col_idx, value=val)
                        cell.border = thin_border
                        if val is not None:
                            cell.alignment = Alignment(horizontal='right')
                    current_row += 1
                
                current_row += 1  # Blank row between sections
            
            # Section 1: Actuals
            def get_actuals(mp):
                if metric not in all_data or mp not in all_data[metric]:
                    return None
                mp_data = all_data[metric][mp]
                if not isinstance(mp_data, dict):
                    return None
                weeks = mp_data.get('weeks', [])
                values = mp_data.get('values', [])
                return {w: values[i] for i, w in enumerate(weeks) if i < len(values)}
            
            write_section('Actuals', get_actuals, section_fills['Actuals'])
            
            # Section 2: Manual FC
            if manual_fc_data and metric in manual_fc_data:
                def get_manual_fc(mp):
                    if mp not in manual_fc_data.get(metric, {}):
                        return None
                    mp_data = manual_fc_data[metric][mp]
                    if not isinstance(mp_data, dict):
                        return None
                    weeks = mp_data.get('weeks', [])
                    values = mp_data.get('values', [])
                    return {w: values[i] for i, w in enumerate(weeks) if i < len(values)}
                
                write_section('Manual FC', get_manual_fc, section_fills['Manual FC'])
            
            # Section 3: Model FC (baseline)
            if metric in model_forecasts and model_forecasts[metric]:
                def get_model_fc(mp):
                    fc = model_forecasts.get(metric, {}).get(mp)
                    if not fc or 'dates' not in fc:
                        return None
                    weeks = dates_to_weeks(fc['dates'])
                    values = fc['values']
                    return {w: round(values[i], 4 if metric == 'Transit Conversion' else 2) 
                            for i, w in enumerate(weeks) if i < len(values)}
                
                write_section('Model FC', get_model_fc, section_fills['Model FC'])
            
            # Section 4: Model FC +Promo
            if data_processor.has_promo_scores and metric in model_promo_forecasts and model_promo_forecasts[metric]:
                def get_model_promo_fc(mp):
                    fc = model_promo_forecasts.get(metric, {}).get(mp)
                    if not fc or 'dates' not in fc:
                        return None
                    weeks = dates_to_weeks(fc['dates'])
                    values = fc['values']
                    return {w: round(values[i], 4 if metric == 'Transit Conversion' else 2) 
                            for i, w in enumerate(weeks) if i < len(values)}
                
                write_section('Model FC +Promo', get_model_promo_fc, section_fills['Model FC +Promo'])
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.now().strftime('%Y-%m-%d')
        filename = f'amazon_haul_eu5_export_{timestamp}.xlsx'
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
        
    except ImportError:
        return jsonify({
            'success': False, 
            'error': 'openpyxl library not installed. Please install it with: pip install openpyxl'
        }), 500
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# Static file serving for development
@app.route('/static/<path:filename>')
def serve_static(filename):
    """Serve static files"""
    return send_from_directory('static', filename)


def auto_load_data():
    """Try to auto-load inputs file from SharePoint, then local fallback."""
    global data_processor, current_file
    
    # Try SharePoint first
    if os.path.exists(SHAREPOINT_INPUTS_FILE):
        print(f"  → Found inputs file on SharePoint")
        try:
            data_processor = DataProcessor()
            data_processor._backtest_cache = {}
            success, message = data_processor.load_excel(SHAREPOINT_INPUTS_FILE)
            if success:
                data_processor.calculate_eu5_totals()
                current_file = "inputs_forecasting.xlsx (SharePoint)"
                print(f"  ✓ Auto-loaded from SharePoint: {message}")
                return True
            else:
                print(f"  ✗ SharePoint file load failed: {message}")
                data_processor = None
        except Exception as e:
            print(f"  ✗ SharePoint load error: {e}")
            data_processor = None
    else:
        print(f"  → SharePoint not accessible (VPN/offline?)")
    
    # Fallback to local file
    if os.path.exists(LOCAL_INPUTS_FILE):
        print(f"  → Trying local fallback: {LOCAL_INPUTS_FILE}")
        try:
            data_processor = DataProcessor()
            data_processor._backtest_cache = {}
            success, message = data_processor.load_excel(LOCAL_INPUTS_FILE)
            if success:
                data_processor.calculate_eu5_totals()
                current_file = "inputs_forecasting.xlsx (local)"
                print(f"  ✓ Auto-loaded from local: {message}")
                return True
            else:
                print(f"  ✗ Local file load failed: {message}")
                data_processor = None
        except Exception as e:
            print(f"  ✗ Local load error: {e}")
            data_processor = None
    
    print("  ⚠ No input file found. Upload manually via dashboard.")
    return False


if __name__ == '__main__':
    print("\n" + "="*60)
    print("  Amazon Haul EU5 Forecasting Dashboard v" + APP_VERSION)
    print("="*60)
    
    print("\n  Auto-loading data...")
    auto_load_data()
    
    print(f"\n  Starting server at http://localhost:5000")
    print("  Press Ctrl+C to stop the server\n")
    print("="*60 + "\n")
    
    app.run(debug=True, host='0.0.0.0', port=5000)
